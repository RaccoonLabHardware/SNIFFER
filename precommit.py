import json
import os

def getProjectName():
    for path in os.listdir('./'):
        #print(path)
        if '.prjpcb' in path.lower():
            _PROJECT = path
            break
    print(_PROJECT)
    return _PROJECT

PROJECT = getProjectName()

def getParametersFromProjPCB(_PROJECT):
    prj = open('./'+_PROJECT,mode="r",encoding="utf-8")

    parameters = {}

    while prj:
        line = prj.readline()
        if '[Parameter' in line:
            n = prj.readline().split('=')[1].strip()
            v = prj.readline().split('=')[1].strip()
            parameters[n] = v
            print(n,v)
        if line == '':
            break
    prj.close()

    #version in PRJ
    vp_h = str(parameters['Version'].split('.')[0] )
    vp_m = str(parameters['Version'].split('.')[1] if len(parameters['Version'].split('.'))>1 else 0)
    vp_l = str(parameters['Version'].split('.')[2] if len(parameters['Version'].split('.'))>2 else 0)
    vp = vp_h+'.'+vp_m+'.'+vp_l
    parameters['vp'] = vp
    print(f'version from {_PROJECT} file is ',vp)
    return parameters

parameters = getParametersFromProjPCB(PROJECT)
vp = parameters['vp']
#PARSE netlist

def parseNetlist():
    net = os.listdir('./Project Outputs/WireListNetlist/')[0]

    with open(f'./Project Outputs/WireListNetlist/{net}') as f:
        lines = f.readlines()
    for i in range(lines.__len__()):
        lines[i] = lines[i].strip()
    lines = lines[lines.index('<<< Wire List >>>'):]
    while '' in lines:
        lines.remove('')

    keys = ['net','designator','pinNum','pinName','component']
    data = lines[2:]
    for i in range(data.__len__()):
        data[i] = data[i].split()
    result = []
    net = ''
    for i in data:
        if i[0][0] == '[':
            net = i[1]
        else:
            result.append({
                'net':net,
                'designator': i[0],
                'pinNum': i[1],
                'pinName': i[2],
                'component': i[-1],
                })
    return result

netlist = parseNetlist()

#MAKE TRANPARENT pics


#   pip install Pillow
# pip install rembg

from rembg import remove
import numpy as np
import cv2

def alphaAndCut(input_path,output_path):
    src = cv2.imread(input_path)
    tgt = remove(src)
    rows = np.any(tgt, axis=1)
    cols = np.any(tgt, axis=0)
    rmin, rmax = np.where(rows)[0][[0, -1]]
    cmin, cmax = np.where(cols)[0][[0, -1]]
    cv2.imwrite(output_path, tgt[rmin:rmax, cmin:cmax])

alphaAndCut('./doc/view.png','./doc/t-view.png')
alphaAndCut('./doc/view-bottom.png','./doc/t-view-bottom.png')
alphaAndCut('./doc/view-top.png','./doc/t-view-top.png')

# GENERATE PNG OF drw

try:
    # pip install pdf2image

    from pdf2image import convert_from_path
    pages = convert_from_path('./doc/doc.pdf', 500)
    pages[-1].save('./doc/drw.png', 'PNG')

    #Cut('./doc/drw.png','./doc/drw.png')
except:
    print('something wrong with pdf2image')


#GET MODEL DIMENSIONS

# pip install steputils
from steputils import p21

def getBBoxFromSTEP():
    step_file = ''

    for path in os.listdir('./doc'):
        if '.step' in path.lower():
            step_file = path
            break
    
    file = p21.readfile(f'doc/{step_file}')

    points = [
        file.data[0].instances[x].entity.params[1] for x in [                      # 4 get point by id
            entry.entity.params[1] for entry in                                    # 3 get vertex point id
            sum([list(section.instances.values()) for section in file.data], [])   # 1 gather all sections
            if hasattr(entry, 'entity') and entry.entity.name == 'VERTEX_POINT'    # 2 filter all vertices
        ]
    ]

    min_x = min(points, key=lambda x: x[0])[0]
    max_x = max(points, key=lambda x: x[0])[0]

    min_y = min(points, key=lambda x: x[1])[1]
    max_y = max(points, key=lambda x: x[1])[1]

    min_z = min(points, key=lambda x: x[2])[2]
    max_z = max(points, key=lambda x: x[2])[2]

    bbox = (
        (min_x, min_y, min_z),
        (max_x, max_y, max_z)
    )

    dim = (
        max_x - min_x,
        max_y - min_y,
        max_z - min_z
    )

    #print(bbox)
    #print(dim)

    x = round(dim[0]*10)/10.0
    y = round(dim[1]*10)/10.0
    z = round(dim[2]*10)/10.0

    print(x,y,z)
    return [x,y,z]

stepBBox = getBBoxFromSTEP()


# GET GERBER DIMENSIONS

from gerber import load_layer

def getBBoxFromGerber(file_name='PCB.GM2'):
    # Open the gerber files
    gm2 = load_layer(f'./Project Outputs/Gerber/{file_name}')

    #print(gm2.__dict__)
    #print(gm2.bounds)
    print(-gm2.bounds[0][0]+gm2.bounds[0][1])
    print(-gm2.bounds[1][0]+gm2.bounds[1][1])
    gerber_x = round((-gm2.bounds[0][0]+gm2.bounds[0][1])*100)/100
    gerber_y = round((-gm2.bounds[1][0]+gm2.bounds[1][1])*100)/100
    #gm2.close()
    return [gerber_x,gerber_y]

gerberBBox = getBBoxFromGerber()

# GENERATE BOM file exactly for PCBWay

import openpyxl
import pandas as pd

def generatBOMToPCBWay(save_to = './Project Outputs/BOM/output.xlsx'):
    with open('./Project Outputs/BOM/BOMtxt-BOM.txt') as f:
        lines = f.readlines()
    for i in range(lines.__len__()):
        lines[i] = lines[i].strip()
    for i in range(lines.__len__()):
        lines[i] = lines[i].split('\t')
    for i in range(len(lines)):
        for j in range(len(lines[i])):
            lines[i][j] = lines[i][j].replace('"', '')
    data = pd.DataFrame(lines[1:], columns=lines[0])
    data = data.drop(0)
    columns = ['*Designator', '*Qty', 'Manufacturer', '*Mfg Part #', 'Value / Description', '*Package/Footprint',
            'Mounting Type', 'Your Instructions / Notes', 'Assembly', '*Unit Price(XX sets)', '*Total', '*Delivery Time',
            '*Actual Purchase Mfg Part #', '*PCBWay Note', 'Customer Reply', 'PCBWay Update']
    dataxls = pd.DataFrame(columns=columns)
    dataxls['*Designator'] = data['Designator']
    dataxls['*Qty'] = data['Quantity']
    dataxls['Manufacturer'] = data['MF']
    dataxls['*Mfg Part #'] = data['MP']
    dataxls['Value / Description'] = data['Value'].astype('str') + '; ' + data['Description'].astype('str')
    dataxls['*Package/Footprint'] = data['Package']
    dataxls['Mounting Type'] = data['Type']
    dataxls['Your Instructions / Notes'] = data['Instructions'] + '; ' + data['HelpURL']
    dataxls.to_excel(save_to)

    book = openpyxl.load_workbook(save_to)
    sheet = book['Sheet1']
    for col in sheet.columns: # автоматическая ширина ячеек
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length * 0.8
        sheet.column_dimensions[column].width = adjusted_width
    yellow = openpyxl.styles.PatternFill('solid', fgColor="FFFF00")
    grey = openpyxl.styles.PatternFill('solid', fgColor="C0C0C0")
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=1, max_col=sheet.max_column):
        for cell in row:
            cell.fill = grey
    sheet['B1'].fill = yellow
    sheet['C1'].fill = yellow
    sheet['E1'].fill = yellow
    sheet['G1'].fill = yellow
    sheet['K1'].fill = yellow
    sheet['L1'].fill = yellow
    sheet['M1'].fill = yellow
    sheet['N1'].fill = yellow
    sheet['O1'].fill = yellow
    sheet['P1'].fill = yellow
    sheet['Q1'].fill = yellow
    book.save(save_to)
    return data


path = "./PCBWay-output"
isExist = os.path.exists(path)
if not isExist:
   os.makedirs(path)

bom = generatBOMToPCBWay('./PCBWay-output/BOM.xlsx')

import shutil

shutil.copyfile('./Project Outputs/Pick Place/Pick Place for PCB.txt', './PCBWay-output/Pick Place.txt')
# 2nd option
path = "./PCBWay-output/Gerber"
isExist = os.path.exists(path)
if not isExist:
   os.makedirs(path)
for path in os.listdir('./Project Outputs/Gerber/'):
    print(path)
    shutil.copyfile(f'./Project Outputs/Gerber/{path}', f'./PCBWay-output/Gerber/{path}')
#shutil.copy('./Project Outputs/Gerber/', './PCBWay-output/Gerber/')  # dst can be a folder; use shutil.copy2() to preserve timestamp
shutil.copyfile('./Project Outputs/CAMtastic1.Cam', './PCBWay-output/CAMtastic1.Cam')
shutil.copyfile('./doc/drw.png', './PCBWay-output/drw.png')
#PCB.xls

# pip install xlrd

def layerStackParce():
    data = pd.read_excel('./Project Outputs/Report Board Stack/PCB.xls')
    # суммарная толщина всех слоёв
    maxvalue = float(list(data[-1:].stack())[0].split()[-1].replace('mm','').replace(',','.')) # Последняя строка из таблицы стакается в одну ячейку, превращается в список, оставшаяся строка делится по пробелам, удаляется mm, запятая меняется на точку, всё это превращается в число float
    data = data[4:-3]
    layers = data.loc[data.iloc[:,-3] == 'Copper'] #Ищу по столбцу где есть медь - это проводящие слои
    listlayers = list(layers.iloc[:,-4]) # Список слоёв
    hight = list(layers.iloc[:,-2]) # Список толщин
    temp = []
    for i in hight:
        temp.append(float(i.replace('mm','').replace(',','.')))
    hight = temp.copy()
    standrdrow = [0.2, 0.3, 0.4, 0.6, 0.8, 1.0, 1.2, 1.6, 2.0, 2.4, 2.6, 2.8, 3.0, 3.2]
    standrdvalue = [abs(i-maxvalue) for i in standrdrow]
    standrdvalue = standrdrow[standrdvalue.index(min(standrdvalue))]

    re = {}
    re['height'] = hight
    re['list'] = listlayers
    re['stndart height'] = standrdvalue
    return re

layerStack = layerStackParce()
#PCB.TXT

def minFromNCDrill(file_name = './Project Outputs/NC Drill/PCB.TXT'):
    with open(file_name) as f:
        lines = f.readlines()
    for i in range(lines.__len__()):
        lines[i] = lines[i].strip()
    pcbplate = lines[lines.index(';TYPE=PLATED')+1:lines.index('%')] #TODO в цикл его надо пихать шобы по всему файлу пройтись?
    tmp = []
    for i in pcbplate:
        if 'F00S00C' in i:
            tmp.append(float(i[i.index('C')+1:]))
    return min(tmp)

drill_files = []
for path in os.listdir('./Project Outputs/NC Drill/'):
    if '.TXT'.lower() in path.lower():
        drill_files.append('./Project Outputs/NC Drill/'+path)
        shutil.copyfile('./Project Outputs/NC Drill/'+path, './PCBWay-output/'+path)

mindrill = 50
for item in drill_files:
    mindrill = min( minFromNCDrill(item),mindrill)
#PCB.G1 GTL GBL G2 ... G10

def minTrace(file_name = 'PCB.GTL'):
    with open(f'./Project Outputs/Gerber/{file_name}') as f:
        lines = f.readlines()
    for i in range(lines.__len__()):
        lines[i] = lines[i].strip('*\n')
    listfromPCBG1 = []
    for i in lines:
        if i.find('%') == False: #TODO поиск строк по % во всём файле вообще корректен?
            i = i.strip('*%')
            if i.find('C') != -1 and i.find(',') != -1:
                listfromPCBG1.append(float(i[i.find(',')+1:]))

    return min(listfromPCBG1)

minwidth = minTrace()



f = open('README.md', "w")

f.write(f"# {PROJECT.split('.')[0]} v{str(vp)} hardware \n\n")
f.write('| View | Top | Bottom |\n')
f.write('| ---- | --- | ------ |\n')
f.write('| <img src="doc/t-view.png" alt="drawing" width="300"> | <img src="doc/t-view-top.png" alt="drawing" width="300"/> | <img src="doc/t-view-bottom.png" alt="drawing" width="300"/> |\n')
f.write('|  | <img src="doc/r-view-top.jpg" alt="drawing" width="300"/> | <img src="doc/r-view-bottom.jpg" alt="drawing" width="300"/> |\n')
f.write('\n')
f.write('## Features\n\n')

f.write('## Wiring\n\n')

f.write('Schematic features. Schematic can be provided via issue.\n\n')

f.write('**Connectors**\n\n')

f.write(f"The node has connectors which are described in the table below.\n\n")

f.write(f"| N | Connector | Description |\n")
f.write(f"| - | - | - |\n")

i=1
con_des = []
for index, row in bom.iterrows():
    print(row['System'], row['Designator'])
    if 'Connector' in row['System']:
        f.write(f"| {i} | {row['Designator']} |  |\n")
        con_des = con_des+ row['Designator'].replace(' ','').split(',')
        i+=1

f.write('\n[Here](https://docs.raccoonlab.co/guide/wires/) you can find manufacturer part number of connectors it self and its mates.\n\n')

f.write('## Pin configuration and functions\n')
f.write('\n')

pinnames = []
for it in con_des:
    pn = {}
    pn['name'] = it
    #f.write(f"**{it}** \n\n")
    #f.write(f"| Pin N | Net name |\n")
    #f.write(f"| -     | -        |\n")
    
    l=[]
    for item in netlist:
        if it.lower() in item['designator'].lower():
            l.append([item['pinNum'],item['net']])
    l = sorted(l,key=lambda x: (x[0]))
    pn['list'] = l
    #for item in l:
    #    f.write(f"| {item[0]:2} | {item[1]:10} |\n")
    #f.write('\n')
    pinnames.append( pn )

#pinnames = sorted(pinnames,key=lambda x: (len(x['list'])))

strarr = []
strarr.append('|')
strarr.append('|')
for item in pinnames:
    strarr[0] += ' Pin N | '+item['name']+' |'
    strarr[1] += ' ----- | ---------------- |'
    i=3
    for it in item['list']:
        if len(strarr)<i:
            strarr.append('|')
        strarr[i-1] += f' {it[0]} | {it[1]} |'
        i+=1
    while i<len(strarr):
        strarr[i-1] += " | |"
        i+=1
for item in strarr:
    f.write(item+'\n')

f.write('\n\n')
f.write('Here you can see all connections of MCU.\n\n')

f.write('<img src="doc/pinout.png" alt="pinout"/>\n\n')


f.write("| MCU PIN         | PIN Numer | NET Name | Description |\n")
f.write("| ---------- |  -- | --------------  | - |\n")

for item in netlist:
    if 'STM'.lower() in item['component'].lower():
        print(item['pinName'],item['net'])
        f.write(f"| {item['pinName']:14} |  {item['pinNum']:2} | {item['net']:10}  |  |\n")

f.write(f'''

## Specifications

**Mechanical**

Scheme is shown on the picture below. CAN model can be provided via email request or issue on github or downloaded on GrabCAD (opens new window).

<img src="doc/drw.png" alt="drawing" height="400"/>

|       | Width, mm | Length, mm | Height, mm |
| ----- | --------- | ---------- | ---------- |
|Outline| {stepBBox[0]:9} | {stepBBox[1]:10} | {stepBBox[2]:10} |
|PCB    | {gerberBBox[0]:9} | {gerberBBox[1]:10} | {layerStack['stndart height']:10} |

Total weight of device less than 50 g.

### Housing

Information about case presented here.

### Absolute Maximum Ratings

### Recommended operating conditions

### ESD ratings

### MTFF

## Integration

**Recommended mechanical mounting**

**Connection example diagram**

### Power Supply Recommendations

Device is designed to operate from an input voltage supply range between 4.5 V and 5.5 V over CAN2 or CAN3 connector, or 5.5 - 30 V from CAN1. This input supply must be able to withstand the maximum input current and maintain a stable voltage. The resistance of the input supply rail should be low enough that an input current transient does not cause a high enough drop that can cause a false UVLO fault triggering and system reset. The amount of bulk capacitance is not critical, but a 47-uF or 100-uF electrolytic capacitor is a typical choice.

## Revision history

|View |Version| Date| Description|
|-    |-      |-    |-           |



## Order details

### PCB Specification Selection

- Board type : Panel by PCBWay
- Break-away rail: Yes
- Instructions:
~~~
Final size is larger ( {stepBBox[0]} x {stepBBox[1]} mm ) than board it self ( {gerberBBox[0]} x {gerberBBox[1]} mm), 
take a look at the picure in attachements. 
Panel should be designed to be able to install PWM1, PWM2 while assembly.
~~~
- Route Process: Panel as PCBWay prefer
- X-out Allowance in Panel:  Accept

- Size (single): {gerberBBox[0]} x {gerberBBox[1]} mm
- Quantity (single): 200
- Layers: {len(layerStack["list"])} -   {layerStack["list"]} check [PCBway layer stack](https://www.pcbway.com/multi-layer-laminated-structure.html)

- Material: FR-4
- FR4-TG: TG 150-160
- Thickness: {layerStack["stndart height"]}
- Min Track/Spacing: {round(minwidth*39.3701)}/{round(minwidth*39.3701)}mil ({minwidth} mm)
- Min Hole Size: {mindrill} mm
- Solder Mask: Black
- Silkscreen: White
- Edge connector: No
- Surface Finish: HASL with lead
- Yes - Tick means you accept we might change "HASL" to "ENIG" at our discretion without extra charge.
- Via Process: Tenting vias
- Finished Copper: 1 oz Cu
- Other Special request:
~~~
Final size is larger ( {stepBBox[0]} x {stepBBox[1]} mm ) than board it self ( {gerberBBox[0]} x {gerberBBox[1]} mm )
~~~

### Assembly Service

- Turnkey
- Board type : Panelized PCBs
-  Assembly Side(s): Both sides
- Quantity: 200
- Contains Sensitive components/parts - No; 
- Do you accept alternatives/substitutes made in China? - Yes

- Number of Unique Parts: 0
- Number of SMD Parts: 0
- Number of BGA/QFP Parts: 0
- Number of Through-Hole Parts: 0

### Additional Options

- Firmware loading: Yes
- Detailed information of assembly:
~~~
Firmware is in attachements.
Take a look at the picure in attachements should be installed from the side.
~~~

## Device and Documentation Support

- [User manual]()
- [Hardware docs](doc/doc.pdf)

## Device Support

- [Firmware sources]()
- [Firmware binary]()

## TERMS OF USAGE / LICENCE

The material provided in this Github repository is subject to the following conditions. 

Firmware files: All firmwares are free (but not open source). Besides unlimited private use you are also granted the permission to use them for commercial purposes under the condition that (1) you dont modify the firmware, e.g. remove or change copyright statements, (2) provide it for free, i.e. dont charge any explicit or implicit fees to your customers, and (3) correctly and clearly cite the origin of the firmware and the project web page in any product documentation or web page. 

Hardware files: All hardware, for which material is provided, is open source hardware, under the terms of the TAPR Open Hardware License as published by the Free Hardware Foundation, see http://www.tapr.org/ohl.html. The TAPR license explicitly permits essentially unlimited commercial use, with only few conditions such as that copyright logos are not removed.

''')






f.close()